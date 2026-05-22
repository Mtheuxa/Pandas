"""
Extrator de Nomes de Alunos — PDF → Excel
==========================================
Uso:
    python ordem_alpha.py [arquivo.pdf] [saida.xlsx]

Dependências:
    pip install pdfplumber openpyxl
"""

import sys
import re
import unicodedata
from pathlib import Path

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Configuração — ajuste conforme o formato do seu PDF
# ---------------------------------------------------------------------------

MODO = "lista"

COLUNA_NOME = None   # None = detecta automaticamente

PALAVRAS_IGNORAR = {
    "nome", "aluno", "alunos", "estudante", "estudantes",
    "matrícula", "matricula", "turma", "série", "serie",
    "escola", "colégio", "colegio", "universidade", "instituição",
    "nº", "no", "número", "numero", "ordem", "total",
    "professor", "professora", "diretor", "coordenador",
    "data", "assinatura", "lista", "chamada", "presença",
    "situação", "situacao", "status"
}

# ---------------------------------------------------------------------------


def normalizar(texto: str) -> str:
    """Remove acentos para comparações, mantendo o texto original intacto."""
    return unicodedata.normalize("NFD", texto).encode("ascii", "ignore").decode()


def parece_nome(linha: str) -> bool:
    """
    Heurística para identificar se uma linha é um nome de pessoa.
    """
    linha = linha.strip()
    if not linha or len(linha) < 5:
        return False

    if re.search(r"\d", linha):
        return False

    if len(linha) > 100:
        return False

    palavras = linha.split()
    # Aumentado para 10 para suportar nomes maiores
    if not (2 <= len(palavras) <= 10):
        return False

    norm = normalizar(linha.lower())
    if norm in {normalizar(p) for p in PALAVRAS_IGNORAR}:
        return False

    if "curso:" in norm or "classe:" in norm or "serie:" in norm or "série:" in norm:
        return False
    if "semipresencial" in norm:
        return False

    com_maiuscula = sum(1 for p in palavras if p[0].isupper())
    if com_maiuscula < len(palavras) // 3:
        return False

    if re.search(r"[=@#$%&*<>{}\\|]", linha):
        return False

    return True


def extrair_de_tabelas(pdf) -> list[tuple[str, str]]:
    alunos = []
    idx_nome_global = None
    idx_situacao_global = None
    
    for pagina in pdf.pages:
        tabelas = pagina.extract_tables()
        for tabela in tabelas:
            if not tabela:
                continue

            idx_nome = None
            idx_situacao = None
            eh_cabecalho = False
            cabecalho = [str(c).strip().lower() if c else "" for c in tabela[0]]

            if COLUNA_NOME:
                alvo = normalizar(COLUNA_NOME.lower())
                for i, c in enumerate(cabecalho):
                    if normalizar(c) == alvo:
                        idx_nome = i
                        eh_cabecalho = True
                        break
            else:
                for i, c in enumerate(cabecalho):
                    nc = normalizar(c)
                    if "nome" in nc or "aluno" in nc or "estudante" in nc:
                        idx_nome = i
                        eh_cabecalho = True
                    elif "situacao" in nc or "status" in nc or "estado" in nc:
                        idx_situacao = i
                        eh_cabecalho = True

            if eh_cabecalho:
                idx_nome_global = idx_nome
                idx_situacao_global = idx_situacao
                inicio = 1
            else:
                idx_nome = idx_nome_global
                idx_situacao = idx_situacao_global
                inicio = 0

            for linha in tabela[inicio:]:
                if not linha:
                    continue
                if idx_nome is not None:
                    valor_nome = linha[idx_nome] if idx_nome < len(linha) else ""
                    if valor_nome and parece_nome(str(valor_nome).strip()):
                        nome = str(valor_nome).strip()
                        situacao = str(linha[idx_situacao] if idx_situacao is not None and idx_situacao < len(linha) else "").strip()
                        alunos.append((nome, situacao))
                else:
                    for celula in linha:
                        if celula and parece_nome(str(celula).strip()):
                            alunos.append((str(celula).strip(), ""))
    return alunos


def extrair_de_texto(pdf) -> list[tuple[str, str]]:
    alunos = []
    status_conhecidos = r"(ativa|ativo|inativa|inativo|cancelada|cancelado|transferida|transferido|trancada|trancado|matriculada|matriculado)"
    for pagina in pdf.pages:
        texto = pagina.extract_text() or ""
        for linha in texto.splitlines():
            linha_limpa = re.sub(r"^\s*\d+[\.\)\-\s]+", "", linha).strip()
            
            situacao = ""
            match = re.search(r"\s+" + status_conhecidos + r"$", linha_limpa, re.IGNORECASE)
            if match:
                situacao = match.group(1).capitalize()
                linha_limpa = linha_limpa[:match.start()].strip()

            if parece_nome(linha_limpa):
                alunos.append((linha_limpa, situacao))
    return alunos


def extrair_nomes(caminho_pdf: str) -> list[tuple[str, str]]:
    with pdfplumber.open(caminho_pdf) as pdf:
        if MODO == "tabela":
            alunos = extrair_de_tabelas(pdf)
        elif MODO == "lista":
            alunos = extrair_de_texto(pdf)
        else:  # auto
            alunos = extrair_de_tabelas(pdf)
            if len(alunos) < 2:
                alunos = extrair_de_texto(pdf)

    vistos = set()
    unicos = []
    for nome, situacao in alunos:
        chave = normalizar(nome.lower())
        n_limpo = " ".join(nome.split())
        s_limpo = " ".join(situacao.split()) if situacao else ""
        
        s_lower = s_limpo.lower()
        if s_lower and "ativ" not in s_lower:
            continue
            
        if chave not in vistos:
            vistos.add(chave)
            unicos.append((n_limpo, s_limpo))

    # Ordena pelo nome do aluno
    return sorted(unicos, key=lambda a: normalizar(a[0].lower()))


# ---------------------------------------------------------------------------
# Exportação para Excel
# ---------------------------------------------------------------------------

COR_CABECALHO = "1F3864"   # Azul escuro
COR_LINHA_PAR = "DCE6F1"   # Azul muito claro
BORDA = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)


def exportar_excel(alunos: list[tuple[str, str]], caminho_saida: str, origem_pdf: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Turma Única"

    # ── Cabeçalho ──────────────────────────────────────────────────────────
    ws.merge_cells("A1:D1")
    ws["A1"] = f"Lista de Alunos - Turma Única"
    ws["A1"].font = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", start_color=COR_CABECALHO)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # ── Subtítulo (Informações da Turma) ───────────────────────────────────
    ws.merge_cells("A2:D2")
    ws["A2"] = "Curso: Semipresencial - Biomedicina  |  Série: S - 2026.1.1P - Biomedicina  |  Classe: ge Sala"
    ws["A2"].font = Font(name="Arial", bold=True, size=10, color="1F3864")
    ws["A2"].fill = PatternFill("solid", start_color="DCE6F1")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    # ── Subcabeçalho das colunas ───────────────────────────────────────────
    colunas = ["Nº", "Nome do Aluno", "Situação", "Assinatura"]
    for col, titulo in enumerate(colunas, start=1):
        celula = ws.cell(row=3, column=col, value=titulo)
        celula.font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
        celula.fill = PatternFill("solid", start_color="2E75B6")
        celula.alignment = Alignment(horizontal="center", vertical="center")
        celula.border = BORDA
    ws.row_dimensions[3].height = 22

    # ── Dados ──────────────────────────────────────────────────────────────
    for i, (nome, situacao) in enumerate(alunos, start=1):
        linha = i + 3
        cor = COR_LINHA_PAR if i % 2 == 0 else "FFFFFF"
        fill = PatternFill("solid", start_color=cor)

        # Nº
        c_num = ws.cell(row=linha, column=1, value=i)
        c_num.font = Font(name="Arial", size=10)
        c_num.alignment = Alignment(horizontal="center")
        c_num.fill = fill
        c_num.border = BORDA

        # Nome
        c_nome = ws.cell(row=linha, column=2, value=nome)
        c_nome.font = Font(name="Arial", size=10)
        c_nome.alignment = Alignment(horizontal="left", vertical="center")
        c_nome.fill = fill
        c_nome.border = BORDA

        # Situação
        c_sit = ws.cell(row=linha, column=3, value=situacao)
        c_sit.font = Font(name="Arial", size=10)
        c_sit.alignment = Alignment(horizontal="center", vertical="center")
        c_sit.fill = fill
        c_sit.border = BORDA

        # Assinatura (em branco)
        c_ass = ws.cell(row=linha, column=4, value="")
        c_ass.fill = fill
        c_ass.border = BORDA

        ws.row_dimensions[linha].height = 18

    # ── Rodapé ─────────────────────────────────────────────────────────────
    rodape_linha = len(alunos) + 3
    ws.merge_cells(f"A{rodape_linha}:D{rodape_linha}")
    ws[f"A{rodape_linha}"] = f"Total de alunos: {len(alunos)}"
    ws[f"A{rodape_linha}"].font = Font(name="Arial", bold=True, size=10, italic=True)
    ws[f"A{rodape_linha}"].alignment = Alignment(horizontal="right")

    # ── Largura das colunas ────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = max(
        (len(a[0]) for a in alunos), default=20
    ) + 4
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 30

    # ── Congela a linha de cabeçalho ───────────────────────────────────────
    ws.freeze_panes = "A4"

    wb.save(caminho_saida)


# ---------------------------------------------------------------------------
# Ponto de entrada
# ---------------------------------------------------------------------------

def main():
    # Define caminhos padrão para facilitar o uso no seu ambiente
    caminho_pdf_padrao = r"c:\Users\unig.ead\Desktop\Pandas\Ordem.alph\BIO 1P MANHÃ.pdf"
    caminho_saida_padrao = r"c:\Users\unig.ead\Desktop\Pandas\Ordem.alph\Turma_Formatada_v2.xlsx"

    caminho_pdf = sys.argv[1] if len(sys.argv) > 1 else caminho_pdf_padrao
    caminho_saida = sys.argv[2] if len(sys.argv) > 2 else caminho_saida_padrao

    if not Path(caminho_pdf).exists():
        print(f"Erro: arquivo '{caminho_pdf}' não encontrado.")
        sys.exit(1)

    print(f"Lendo: {caminho_pdf}")
    alunos = extrair_nomes(caminho_pdf)

    if not alunos:
        print("Aviso: Nenhum nome de aluno encontrado.")
        print("   Dica: ajuste a variável MODO ou COLUNA_NOME no início do script.")
        sys.exit(1)

    print(f"Encontrado(s) {len(alunos)} aluno(s).")
    exportar_excel(alunos, caminho_saida, caminho_pdf)
    print(f"Planilha salva em: {caminho_saida}")


if __name__ == "__main__":
    main()
